
from openpyxl import Workbook,load_workbook
import datetime

# 创建一个excel
# wb = Workbook()
# # 获取当前active的sheet
# sheet = wb.active
# sheet.title='Alex大魔王的姑娘们...'
#
# # 写数据
# # ⽅式⼀：数据可以直接分配到单元格中(可以输⼊公式)
# sheet['B9']='white girl'
# sheet['C9']='171,48,88'
# # ⽅式⼆：可以附加⾏，从第⼀列开始附加(从最下⽅空⽩处，最左开始)(可以输⼊多⾏)
# sheet.append(['rose','170','49'])
# # ⽅式三：Python 类型会被⾃动转换
# sheet['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
#
# wb.save('excel_test.xlsx')


# 打开已有文件
wb = load_workbook('迭代开发计划.xlsx')

sheet = wb.get_sheet_by_name('SP(5-1)')

#获取指定列的切片数据
# for cell in sheet['b2:b10']:
#     print(cell[0].value)

# 获取表格所有数据
# for row in sheet:
#     for cell in row:
#         print(cell.value,end=",")
#     print()

# A1, A2, A3这样的顺序
# for column in sheet.columns:
#     for cell in column:
#         print(cell.value,end=",")
#     print()

# 遍历指定⾏&列
# for row in sheet.iter_rows(min_row=4,max_row=21,max_col=5):
#     for cell in row:
#         print(cell.value,end=",")
#     print()

from openpyxl.styles import Font, colors, Alignment

# 给单元格设置样式
myFont = Font(name="宋体",size=20,italic=True,color=colors.BLUE)
sheet['b5'].font=myFont
# 设置B1中的数据垂直居中和⽔平居中
sheet['j5'].alignment = Alignment(horizontal='center', vertical='center')
# 第2⾏⾏⾼
sheet.row_dimensions[2].height = 140
# C列列宽
sheet.column_dimensions['C'].width = 130

wb.save('迭代开发计划.xlsx')